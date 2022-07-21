#-------------------------------------------------
# Jose Ignacio LeBlanc
# Christian Caceres
#-------------------------------------------------

# Librerias
import random 
import time

# EXCEL
from openpyxl import Workbook # pip install openpyxl
import openpyxl
from openpyxl.chart import LineChart, Reference 

# Importamos los Algoritmos de Ordenamiento!!
from alg.BubbleSort import bubbleSort
from alg.MergSort import mergeSort
from alg.QuickSort import quickSort
from alg.SelectionSort import selectionSort
from alg.Radix import radixSort
from alg.HeapSort import heapSort

# Creamos la Instancia del Libro
wb = Workbook()
ws = wb.active # activar

ws['A1'] = "N"
ws['B1'] = "Quicksort"
ws['C1'] = "Mergsort"
ws['D1'] = "SelectionSort"
ws['E1'] = "Bubblesort"
ws['F1'] = "Radix"
ws['G1'] = "HeapSort"

# Bechmarck
base = 1000 # Base para N
n = base
num = 2 # Fila (2) de la tabla

while n <= ( base * 10 ) :

    ws[ 'A' + str(num) ] = n

    ltr = 66 # Columna B
    
    # Inicializacion de array[n] con numeros aleatorios entre (1, 10000000)
    arr = []
    for i in range( 0, n ):
        arr.append( random.randint( 1, 10000000 ) ) # 1 - 10.000.000

    print("-----------------------------------------------\nN = ", n)
    
    # Quicksort
    start_timer = time.time() # Inicio
    quickSort(arr, 0, n - 1) # Funcion
    end_timer = time.time() # Fin

    execution_timer = end_timer - start_timer # Tiempo de Ejecucion
    print("Quicksort\tTiempo : ", ( execution_timer ) )
    ws[ chr(ltr) + str(num) ] = execution_timer # insertar tiempo
    ltr = ltr + 1

    # Mergsort
    start_timer = time.time() # Inicio
    mergeSort(arr, 0, n - 1) # Funcion
    end_timer = time.time() # Fin

    execution_timer = end_timer - start_timer # Tiempo de Ejecucion
    print("Mergsort\tTiempo : ", ( execution_timer ) )
    ws[ chr(ltr) + str(num) ] = execution_timer # insertar tiempo
    ltr = ltr + 1

    # SelectionSort
    if i <= 100000 :

        start_timer = time.time() # Inicio
        selectionSort( arr ) # Funcion
        end_timer = time.time() # Fin

        execution_timer = end_timer - start_timer # Tiempo de Ejecucion
        print("SelectionSort\tTiempo : ", ( execution_timer ) )
        ws[ chr(ltr) + str(num) ] = execution_timer # insertar tiempo
        ltr = ltr + 1

    # Bubblesort
    start_time = time.time() # Inicio
    bubbleSort( arr ) # Funcion
    end_timer = time.time() # Fin

    execution_timer = end_timer - start_timer # Tiempo de Ejecucion
    print("Bubblesort\tTiempo : ", ( execution_timer ) )
    ws[ chr(ltr) + str(num) ] = execution_timer # insertar tiempo
    ltr = ltr + 1

    # Radix
    start_time = time.time() # Inicio
    radixSort( arr ) # Funcion
    end_timer = time.time() # Fin

    execution_timer = end_timer - start_timer # Tiempo de Ejecucion
    print("Radix\t\tTiempo : ", ( execution_timer ) )
    ws[ chr(ltr) + str(num) ] = execution_timer # insertar tiempo
    ltr = ltr + 1

    # HeapSort
    start_time = time.time() # Inicio
    heapSort( arr ) # Funcion
    end_timer = time.time() # Fin

    execution_timer = end_timer - start_timer # Tiempo de Ejecucion
    print("HeapSort\tTiempo : ", ( execution_timer ) )
    ws[ chr(ltr) + str(num) ] = execution_timer # insertar tiempo

    print("-----------------------------------------------")

    num = num + 1
    n = n + base

# Graficar
chart = LineChart()

# Titulos
chart.title = "GRAFICO"
chart.y_axis.title = 'Tiempo'
chart.x_axis.title = 'N elementos'

# Datos de Referencias
datos = Reference ( 
    ws,
    min_col = 2,
    min_row = 1, 
    max_col = 7,
    max_row = 11
)
chart.add_data( datos, titles_from_data = True )

# Categorias de Referencia
cat = Reference (
    ws,
    min_col = 1,
    min_row = 2,
    max_row = 11
) 
chart.set_categories( cat )

# Posicion del Grafico
ws.add_chart( chart, "A13" ) 

# Guardar Libro
wb.save("Grafico.xlsx")